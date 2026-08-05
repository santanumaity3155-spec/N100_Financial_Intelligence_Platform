"""
valuation.py

Valuation Module for the N100 Financial Intelligence Platform.

This module provides functions to calculate valuation metrics for all companies,
compute sector median PE ratios, assign valuation flags, and generate output files.

Functions:
- calculate_fcf_yield: Calculate Free Cash Flow Yield
- calculate_sector_median_pe: Compute median PE by Broad Sector
- calculate_sector_relative_pe: Calculate Company PE vs Sector Median
- assign_valuation_flag: Assign Caution/Discount/Fair flags
- build_valuation_dataframe: Build complete valuation dataframe
- export_valuation_summary: Export to Excel with formatting
- export_valuation_flags: Export flags to CSV
"""

import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config.constants import OUTPUT_DIR, DATABASE_DIR
from src.config.logging_config import get_logger
from src.database.connection import get_connection

logger = get_logger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

# Output paths
VALUATION_SUMMARY_XLSX = OUTPUT_DIR / "valuation_summary.xlsx"
VALUATION_FLAGS_CSV = OUTPUT_DIR / "valuation_flags.csv"

# Valuation flag thresholds
PE_CAUTION_THRESHOLD = 1.5  # PE > Sector Median * 1.5 = Caution
PE_DISCOUNT_THRESHOLD = 0.7  # PE < Sector Median * 0.7 = Discount

# Excel formatting
EXCEL_HEADER_FONT = Font(bold=True, color="FFFFFF")
EXCEL_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
EXCEL_FAIR_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
EXCEL_DISCOUNT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
EXCEL_CAUTION_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _safe_numeric(value: any, default: Optional[float] = None) -> Optional[float]:
    """
    Safely convert value to numeric, handling None, NaN, and non-numeric values.
    
    Parameters
    ----------
    value : any
        Value to convert
    default : Optional[float], optional
        Default value if conversion fails, by default None
    
    Returns
    -------
    Optional[float]
        Numeric value or default
    """
    if value is None:
        return default
    if pd.isna(value):
        return default
    try:
        numeric = float(value)
        return numeric
    except (ValueError, TypeError):
        return default


def _get_latest_period(conn: sqlite3.Connection, table: str, company_id: str) -> Optional[str]:
    """
    Get the latest period for a company from a table.
    
    Parameters
    ----------
    conn : sqlite3.Connection
        Database connection
    table : str
        Table name
    company_id : str
        Company identifier
    
    Returns
    -------
    Optional[str]
        Latest period or None
    """
    try:
        query = f"""
            SELECT MAX(period) as latest_period 
            FROM {table} 
            WHERE company_id = ?
        """
        cursor = conn.execute(query, (company_id,))
        result = cursor.fetchone()
        return result['latest_period'] if result else None
    except Exception as e:
        logger.warning(f"Failed to get latest period for {company_id} from {table}: {str(e)}")
        return None


# =============================================================================
# CORE VALUATION FUNCTIONS
# =============================================================================

def calculate_fcf_yield(free_cash_flow: Optional[float], market_cap: Optional[float]) -> Optional[float]:
    """
    Calculate Free Cash Flow Yield.
    
    Formula: FCF Yield = (Free Cash Flow / Market Cap) × 100
    
    Parameters
    ----------
    free_cash_flow : Optional[float]
        Free Cash Flow value in Crore
    market_cap : Optional[float]
        Market Cap value in Crore
    
    Returns
    -------
    Optional[float]
        FCF Yield percentage, or None if Market Cap is 0 or data missing
    
    Examples
    --------
    >>> calculate_fcf_yield(100, 1000)
    10.0
    >>> calculate_fcf_yield(100, 0)
    None
    >>> calculate_fcf_yield(None, 1000)
    None
    """
    try:
        # Handle None or NaN values
        fcf = _safe_numeric(free_cash_flow)
        mc = _safe_numeric(market_cap)
        
        if fcf is None or mc is None:
            logger.debug("FCF Yield calculation: Missing FCF or Market Cap data")
            return None
        
        # Handle zero market cap
        if mc == 0:
            logger.warning("FCF Yield calculation: Market Cap is zero")
            return None
        
        # Calculate FCF Yield
        fcf_yield = (fcf / mc) * 100
        
        logger.debug(f"FCF Yield calculated: {fcf_yield:.2f}% (FCF={fcf}, MC={mc})")
        return round(fcf_yield, 2)
        
    except Exception as e:
        logger.error(f"FCF Yield calculation failed: {str(e)}")
        return None


def calculate_sector_median_pe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate median PE ratio for each Broad Sector.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with columns: 'broad_sector', 'pe_ratio'
    
    Returns
    -------
    pd.DataFrame
        DataFrame with columns: 'broad_sector', 'sector_median_pe'
    """
    try:
        if df.empty or 'broad_sector' not in df.columns or 'pe_ratio' not in df.columns:
            logger.warning("Sector Median PE calculation: Invalid dataframe provided")
            return pd.DataFrame(columns=['broad_sector', 'sector_median_pe'])
        
        # Filter valid PE ratios (non-null, positive)
        valid_pe = df[df['pe_ratio'].notna() & (df['pe_ratio'] > 0)].copy()
        
        if valid_pe.empty:
            logger.warning("Sector Median PE calculation: No valid PE ratios found")
            return pd.DataFrame(columns=['broad_sector', 'sector_median_pe'])
        
        # Calculate median PE by sector
        sector_medians = valid_pe.groupby('broad_sector')['pe_ratio'].median().reset_index()
        sector_medians.columns = ['broad_sector', 'sector_median_pe']
        
        # Round to 2 decimal places
        sector_medians['sector_median_pe'] = sector_medians['sector_median_pe'].round(2)
        
        logger.info(f"Calculated sector median PE for {len(sector_medians)} sectors")
        return sector_medians
        
    except Exception as e:
        logger.error(f"Sector Median PE calculation failed: {str(e)}")
        return pd.DataFrame(columns=['broad_sector', 'sector_median_pe'])


def calculate_sector_relative_pe(company_pe: Optional[float], sector_median_pe: Optional[float]) -> Optional[float]:
    """
    Calculate Company PE relative to Sector Median PE.
    
    Formula: Sector Relative PE = Company PE / Sector Median PE
    
    Parameters
    ----------
    company_pe : Optional[float]
        Company's PE ratio
    sector_median_pe : Optional[float]
        Sector median PE ratio
    
    Returns
    -------
    Optional[float]
        Sector Relative PE ratio, or None if data missing or sector median is 0
    """
    try:
        # Handle None or NaN values
        comp_pe = _safe_numeric(company_pe)
        sec_med_pe = _safe_numeric(sector_median_pe)
        
        if comp_pe is None or sec_med_pe is None:
            logger.debug("Sector Relative PE calculation: Missing PE or Sector Median PE")
            return None
        
        # Handle zero sector median
        if sec_med_pe == 0:
            logger.warning("Sector Relative PE calculation: Sector Median PE is zero")
            return None
        
        # Calculate relative PE
        relative_pe = comp_pe / sec_med_pe
        
        logger.debug(f"Sector Relative PE calculated: {relative_pe:.2f} (Company PE={comp_pe}, Sector Median={sec_med_pe})")
        return round(relative_pe, 2)
        
    except Exception as e:
        logger.error(f"Sector Relative PE calculation failed: {str(e)}")
        return None


def assign_valuation_flag(pe_vs_sector_median_pct: Optional[float]) -> str:
    """
    Assign valuation flag based on PE vs Sector Median percentage.
    
    Rules:
    - PE > Sector Median × 1.5 (i.e., > 150%) → Flag: Caution
    - PE < Sector Median × 0.7 (i.e., < 70%) → Flag: Discount
    - Else → Flag: Fair
    
    Parameters
    ----------
    pe_vs_sector_median_pct : Optional[float]
        PE vs Sector Median percentage (e.g., 120.0 means 120% of sector median)
    
    Returns
    -------
    str
        Valuation flag: 'Caution', 'Discount', or 'Fair'
    """
    try:
        # Handle None or NaN values
        pe_pct = _safe_numeric(pe_vs_sector_median_pct)
        
        if pe_pct is None:
            logger.debug("Valuation Flag assignment: Missing PE vs Sector Median data")
            return "Fair"  # Default to Fair if data missing
        
        # Apply rules
        if pe_pct > (PE_CAUTION_THRESHOLD * 100):  # > 150%
            flag = "Caution"
        elif pe_pct < (PE_DISCOUNT_THRESHOLD * 100):  # < 70%
            flag = "Discount"
        else:
            flag = "Fair"
        
        logger.debug(f"Valuation Flag assigned: {flag} (PE vs Sector Median={pe_pct}%)")
        return flag
        
    except Exception as e:
        logger.error(f"Valuation Flag assignment failed: {str(e)}")
        return "Fair"


def build_valuation_dataframe() -> pd.DataFrame:
    """
    Build complete valuation dataframe for all companies.
    
    Returns
    -------
    pd.DataFrame
        DataFrame with columns:
        - Company ID
        - Company Name
        - Ticker
        - Sector
        - Sub-sector
        - Broad Sector
        - Market Cap
        - PE
        - PB
        - EV/EBITDA
        - Free Cash Flow
        - FCF Yield %
        - Sector Median PE
        - PE vs Sector Median %
        - Valuation Flag
    """
    start_time = time.time()
    logger.info("=" * 80)
    logger.info("BUILDING VALUATION DATAFRAME")
    logger.info("=" * 80)
    
    try:
        # Step 1: Load database
        logger.info("Step 1: Loading database")
        conn = get_connection()
        logger.info("Database loaded successfully")
        
        # Step 2: Load all required data
        logger.info("Step 2: Loading company and financial data")
        
        # Load companies
        companies_query = """
            SELECT 
                c.company_id,
                c.company_name,
                c.sector,
                s.broad_sector,
                s.sub_sector
            FROM companies c
            LEFT JOIN sectors s ON c.company_id = s.company_id
        """
        companies_df = pd.read_sql_query(companies_query, conn)
        logger.info(f"Loaded {len(companies_df)} companies")
        
        # Load latest market cap data for all companies
        market_cap_query = """
            SELECT 
                mc.company_id,
                mc.market_cap,
                mc.pe_ratio,
                mc.pb_ratio,
                mc.ev_ebitda
            FROM market_cap mc
            INNER JOIN (
                SELECT company_id, MAX(period) as latest_period
                FROM market_cap
                GROUP BY company_id
            ) latest ON mc.company_id = latest.company_id 
            AND mc.period = latest.latest_period
        """
        market_cap_df = pd.read_sql_query(market_cap_query, conn)
        logger.info(f"Loaded market cap data for {len(market_cap_df)} companies")
        
        # Load latest cash flow data (for FCF)
        cash_flow_query = """
            SELECT 
                cf.company_id,
                cf.free_cash_flow
            FROM cash_flow cf
            INNER JOIN (
                SELECT company_id, MAX(period) as latest_period
                FROM cash_flow
                GROUP BY company_id
            ) latest ON cf.company_id = latest.company_id 
            AND cf.period = latest.latest_period
        """
        cash_flow_df = pd.read_sql_query(cash_flow_query, conn)
        logger.info(f"Loaded cash flow data for {len(cash_flow_df)} companies")
        
        # Step 3: Merge all data
        logger.info("Step 3: Merging data")
        valuation_df = companies_df.merge(market_cap_df, on='company_id', how='left')
        valuation_df = valuation_df.merge(cash_flow_df, on='company_id', how='left')
        
        # Fill missing broad_sector with 'Unknown' to enable sector median calculation
        valuation_df['broad_sector'] = valuation_df['broad_sector'].fillna('Unknown')
        valuation_df['sector'] = valuation_df['sector'].fillna('Unknown')
        
        logger.info(f"Companies processed: {len(valuation_df)}")
        
        # Step 4: Calculate FCF Yield (before renaming, using original column names)
        logger.info("Step 4: Calculating FCF Yield")
        valuation_df['fcf_yield_temp'] = valuation_df.apply(
            lambda row: calculate_fcf_yield(row['free_cash_flow'], row['market_cap']),
            axis=1
        )
        
        # Step 5: Calculate Sector Median PE (using original column names)
        logger.info("Step 5: Computing sector median PE")
        sector_medians = calculate_sector_median_pe(valuation_df)
        
        # Merge sector medians
        valuation_df = valuation_df.merge(sector_medians, on='broad_sector', how='left')
        logger.info("Sector medians computed")
        
        # Step 6: Calculate PE vs Sector Median %
        logger.info("Step 6: Calculating PE vs Sector Median %")
        valuation_df['pe_vs_sector_median_pct'] = valuation_df.apply(
            lambda row: round((row['pe_ratio'] / row['sector_median_pe']) * 100, 2)
            if pd.notna(row['pe_ratio']) and pd.notna(row['sector_median_pe']) and row['sector_median_pe'] > 0
            else None,
            axis=1
        )
        
        # Step 7: Assign Valuation Flags
        logger.info("Step 7: Assigning valuation flags")
        valuation_df['valuation_flag'] = valuation_df['pe_vs_sector_median_pct'].apply(assign_valuation_flag)
        logger.info("Valuation flags assigned")
        
        # Step 8: Rename columns to match requirements
        logger.info("Step 8: Preparing final dataframe")
        valuation_df = valuation_df.rename(columns={
            'company_id': 'Company ID',
            'company_name': 'Company Name',
            'sector': 'Sector',
            'broad_sector': 'Broad Sector',
            'sub_sector': 'Sub-sector',
            'market_cap': 'Market Cap',
            'pe_ratio': 'PE',
            'pb_ratio': 'PB',
            'ev_ebitda': 'EV/EBITDA',
            'free_cash_flow': 'Free Cash Flow',
            'fcf_yield_temp': 'FCF Yield %',
            'sector_median_pe': 'Sector Median PE',
            'pe_vs_sector_median_pct': 'PE vs Sector Median %',
            'valuation_flag': 'Valuation Flag'
        })
        
        # Add Ticker (using company_id as ticker for now)
        valuation_df['Ticker'] = valuation_df['Company ID']
        
        # Select and reorder final columns
        final_columns = [
            'Company ID',
            'Company Name',
            'Ticker',
            'Sector',
            'Sub-sector',
            'Broad Sector',
            'Market Cap',
            'PE',
            'PB',
            'EV/EBITDA',
            'Free Cash Flow',
            'FCF Yield %',
            'Sector Median PE',
            'PE vs Sector Median %',
            'Valuation Flag'
        ]
        
        # Only include columns that exist
        available_columns = [col for col in final_columns if col in valuation_df.columns]
        valuation_df = valuation_df[available_columns]
        
        # Sort by Broad Sector and Company Name
        valuation_df = valuation_df.sort_values(['Broad Sector', 'Company Name'], ascending=[True, True])
        
        # Reset index
        valuation_df = valuation_df.reset_index(drop=True)
        
        elapsed_time = time.time() - start_time
        logger.info(f"Valuation dataframe built successfully in {elapsed_time:.2f}s")
        logger.info(f"Total companies: {len(valuation_df)}")
        
        return valuation_df
        
    except Exception as e:
        logger.error(f"Failed to build valuation dataframe: {str(e)}", exc_info=True)
        return pd.DataFrame()


def export_valuation_summary(df: pd.DataFrame, output_path: Path = VALUATION_SUMMARY_XLSX) -> Path:
    """
    Export valuation summary to Excel with formatting.
    
    Parameters
    ----------
    df : pd.DataFrame
        Valuation dataframe
    output_path : Path, optional
        Output file path, by default VALUATION_SUMMARY_XLSX
    
    Returns
    -------
    Path
        Path to exported Excel file
    """
    start_time = time.time()
    logger.info("=" * 80)
    logger.info("EXPORTING VALUATION SUMMARY TO EXCEL")
    logger.info("=" * 80)
    
    try:
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Valuation Summary"
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = EXCEL_HEADER_FONT
            cell.fill = EXCEL_HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply conditional formatting based on Valuation Flag
                if headers[col_idx - 1] == 'Valuation Flag':
                    if value == 'Fair':
                        cell.fill = EXCEL_FAIR_FILL
                    elif value == 'Discount':
                        cell.fill = EXCEL_DISCOUNT_FILL
                    elif value == 'Caution':
                        cell.fill = EXCEL_CAUTION_FILL
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_idx in range(2, min(len(df) + 2, 100)):  # Check first 100 rows for performance
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Enable filter
        ws.auto_filter.ref = ws.dimensions
        
        # Save workbook
        wb.save(output_path)
        
        elapsed_time = time.time() - start_time
        logger.info(f"Excel exported successfully to {output_path} in {elapsed_time:.2f}s")
        logger.info(f"Total rows: {len(df)}, Columns: {len(headers)}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Failed to export Excel: {str(e)}", exc_info=True)
        raise


def export_valuation_flags(df: pd.DataFrame, output_path: Path = VALUATION_FLAGS_CSV) -> Path:
    """
    Export valuation flags to CSV (only Caution and Discount companies).
    
    Parameters
    ----------
    df : pd.DataFrame
        Valuation dataframe
    output_path : Path, optional
        Output file path, by default VALUATION_FLAGS_CSV
    
    Returns
    -------
    Path
        Path to exported CSV file
    """
    start_time = time.time()
    logger.info("=" * 80)
    logger.info("EXPORTING VALUATION FLAGS TO CSV")
    logger.info("=" * 80)
    
    try:
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Filter only Caution and Discount companies
        flags_df = df[df['Valuation Flag'].isin(['Caution', 'Discount'])].copy()
        
        if flags_df.empty:
            logger.warning("No Caution or Discount companies found")
        
        # Select required columns
        required_columns = ['Company Name', 'Ticker', 'Sector', 'PE', 'Sector Median PE', 'PE vs Sector Median %', 'Valuation Flag']
        available_columns = [col for col in required_columns if col in flags_df.columns]
        
        if not available_columns:
            logger.error("No required columns found in dataframe")
            raise ValueError("Missing required columns for CSV export")
        
        flags_df = flags_df[available_columns]
        
        # Calculate difference percentage
        if 'PE' in flags_df.columns and 'Sector Median PE' in flags_df.columns:
            flags_df['Difference %'] = flags_df.apply(
                lambda row: round(((row['PE'] - row['Sector Median PE']) / row['Sector Median PE']) * 100, 2)
                if pd.notna(row['PE']) and pd.notna(row['Sector Median PE']) and row['Sector Median PE'] > 0
                else None,
                axis=1
            )
        
        # Sort by Flag and Sector
        flags_df = flags_df.sort_values(['Valuation Flag', 'Sector'], ascending=[True, True])
        
        # Export to CSV with UTF-8 encoding
        flags_df.to_csv(output_path, index=False, encoding='utf-8')
        
        elapsed_time = time.time() - start_time
        logger.info(f"CSV exported successfully to {output_path} in {elapsed_time:.2f}s")
        logger.info(f"Total companies with flags: {len(flags_df)}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Failed to export CSV: {str(e)}", exc_info=True)
        raise


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_valuation_pipeline() -> Dict[str, any]:
    """
    Run the complete valuation pipeline.
    
    Returns
    -------
    Dict[str, any]
        Pipeline statistics and results
    """
    start_time = time.time()
    logger.info("=" * 80)
    logger.info("STARTING VALUATION PIPELINE")
    logger.info("=" * 80)
    
    stats = {
        "start_time": datetime.now().isoformat(),
        "end_time": None,
        "status": "running",
        "companies_processed": 0,
        "excel_exported": False,
        "csv_exported": False,
        "errors": [],
        "warnings": []
    }
    
    try:
        # Step 1: Build valuation dataframe
        logger.info("Step 1: Building valuation dataframe")
        valuation_df = build_valuation_dataframe()
        
        if valuation_df.empty:
            error_msg = "Valuation dataframe is empty"
            logger.error(error_msg)
            stats["errors"].append(error_msg)
            stats["status"] = "failed"
            return stats
        
        stats["companies_processed"] = len(valuation_df)
        
        # Step 2: Export to Excel
        logger.info("Step 2: Exporting to Excel")
        try:
            export_valuation_summary(valuation_df)
            stats["excel_exported"] = True
            logger.info("Excel export completed")
        except Exception as e:
            error_msg = f"Excel export failed: {str(e)}"
            logger.error(error_msg)
            stats["errors"].append(error_msg)
        
        # Step 3: Export flags to CSV
        logger.info("Step 3: Exporting flags to CSV")
        try:
            export_valuation_flags(valuation_df)
            stats["csv_exported"] = True
            logger.info("CSV export completed")
        except Exception as e:
            error_msg = f"CSV export failed: {str(e)}"
            logger.error(error_msg)
            stats["errors"].append(error_msg)
        
        # Update stats
        elapsed_time = time.time() - start_time
        stats["end_time"] = datetime.now().isoformat()
        stats["total_time_seconds"] = round(elapsed_time, 2)
        stats["status"] = "completed" if not stats["errors"] else "completed_with_errors"
        
        logger.info("=" * 80)
        logger.info("VALUATION PIPELINE COMPLETED")
        logger.info("=" * 80)
        logger.info(f"Companies processed: {stats['companies_processed']}")
        logger.info(f"Excel exported: {stats['excel_exported']}")
        logger.info(f"CSV exported: {stats['csv_exported']}")
        logger.info(f"Total time: {elapsed_time:.2f}s")
        logger.info(f"Status: {stats['status']}")
        
        if stats["errors"]:
            logger.warning(f"Errors encountered: {len(stats['errors'])}")
            for error in stats["errors"]:
                logger.warning(f"  - {error}")
        
        return stats
        
    except Exception as e:
        elapsed_time = time.time() - start_time
        error_msg = f"Pipeline failed: {str(e)}"
        logger.error(error_msg, exc_info=True)
        stats["end_time"] = datetime.now().isoformat()
        stats["total_time_seconds"] = round(elapsed_time, 2)
        stats["status"] = "failed"
        stats["errors"].append(error_msg)
        return stats


# =============================================================================
# CLI ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    # Run the valuation pipeline
    stats = run_valuation_pipeline()
    
    # Print summary
    print("\n" + "=" * 80)
    print("VALUATION PIPELINE SUMMARY")
    print("=" * 80)
    print(f"Status: {stats['status']}")
    print(f"Companies Processed: {stats['companies_processed']}")
    print(f"Excel Exported: {stats['excel_exported']}")
    print(f"CSV Exported: {stats['csv_exported']}")
    print(f"Total Time: {stats.get('total_time_seconds', 0)}s")
    
    if stats['errors']:
        print(f"\nErrors: {len(stats['errors'])}")
        for error in stats['errors']:
            print(f"  - {error}")
    
    print("=" * 80)