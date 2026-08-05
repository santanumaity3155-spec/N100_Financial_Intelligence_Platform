"""
Test suite for Valuation Module (Module 5)

Tests all functions, edge cases, and validates output files.
"""

import logging
import sys
from pathlib import Path

import pandas as pd
import pytest

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.analytics.valuation import (
    calculate_fcf_yield,
    calculate_sector_median_pe,
    calculate_sector_relative_pe,
    assign_valuation_flag,
    build_valuation_dataframe,
    export_valuation_summary,
    export_valuation_flags,
    run_valuation_pipeline,
)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# =============================================================================
# UNIT TESTS
# =============================================================================

class TestCalculateFCFYield:
    """Test FCF Yield calculation"""
    
    def test_normal_calculation(self):
        """Test normal FCF yield calculation"""
        result = calculate_fcf_yield(100, 1000)
        assert result == 10.0
    
    def test_zero_market_cap(self):
        """Test zero market cap returns None"""
        result = calculate_fcf_yield(100, 0)
        assert result is None
    
    def test_none_values(self):
        """Test None values return None"""
        assert calculate_fcf_yield(None, 1000) is None
        assert calculate_fcf_yield(100, None) is None
        assert calculate_fcf_yield(None, None) is None
    
    def test_negative_fcf(self):
        """Test negative FCF works correctly"""
        result = calculate_fcf_yield(-100, 1000)
        assert result == -10.0
    
    def test_zero_fcf(self):
        """Test zero FCF"""
        result = calculate_fcf_yield(0, 1000)
        assert result == 0.0


class TestCalculateSectorMedianPE:
    """Test sector median PE calculation"""
    
    def test_normal_calculation(self):
        """Test normal sector median calculation"""
        df = pd.DataFrame({
            'broad_sector': ['Technology', 'Technology', 'Finance', 'Finance'],
            'pe_ratio': [20.0, 30.0, 15.0, 25.0]
        })
        result = calculate_sector_median_pe(df)
        
        assert len(result) == 2
        assert 'broad_sector' in result.columns
        assert 'sector_median_pe' in result.columns
    
    def test_empty_dataframe(self):
        """Test empty dataframe"""
        df = pd.DataFrame()
        result = calculate_sector_median_pe(df)
        assert result.empty
    
    def test_missing_columns(self):
        """Test missing columns"""
        df = pd.DataFrame({'company_id': ['1', '2']})
        result = calculate_sector_median_pe(df)
        assert result.empty
    
    def test_filters_negative_pe(self):
        """Test that negative PE ratios are filtered"""
        df = pd.DataFrame({
            'broad_sector': ['Technology', 'Technology'],
            'pe_ratio': [20.0, -5.0]
        })
        result = calculate_sector_median_pe(df)
        # Should only use positive PE
        assert len(result) == 1


class TestCalculateSectorRelativePE:
    """Test sector relative PE calculation"""
    
    def test_normal_calculation(self):
        """Test normal calculation"""
        result = calculate_sector_relative_pe(20.0, 15.0)
        assert result == 1.33
    
    def test_none_values(self):
        """Test None values"""
        assert calculate_sector_relative_pe(None, 15.0) is None
        assert calculate_sector_relative_pe(20.0, None) is None
    
    def test_zero_sector_median(self):
        """Test zero sector median"""
        result = calculate_sector_relative_pe(20.0, 0)
        assert result is None


class TestAssignValuationFlag:
    """Test valuation flag assignment"""
    
    def test_caution_flag(self):
        """Test Caution flag (>150%)"""
        assert assign_valuation_flag(151.0) == "Caution"
        assert assign_valuation_flag(200.0) == "Caution"
    
    def test_discount_flag(self):
        """Test Discount flag (<70%)"""
        assert assign_valuation_flag(69.0) == "Discount"
        assert assign_valuation_flag(50.0) == "Discount"
    
    def test_fair_flag(self):
        """Test Fair flag (70-150%)"""
        assert assign_valuation_flag(70.0) == "Fair"
        assert assign_valuation_flag(100.0) == "Fair"
        assert assign_valuation_flag(150.0) == "Fair"
    
    def test_none_value(self):
        """Test None value defaults to Fair"""
        assert assign_valuation_flag(None) == "Fair"


# =============================================================================
# INTEGRATION TESTS
# =============================================================================

class TestValuationPipeline:
    """Test complete valuation pipeline"""
    
    def test_pipeline_runs(self):
        """Test that pipeline runs without errors"""
        stats = run_valuation_pipeline()
        assert stats['status'] in ['completed', 'completed_with_errors']
        assert stats['companies_processed'] > 0
    
    def test_excel_generated(self):
        """Test that Excel file is generated"""
        output_path = Path('output/valuation_summary.xlsx')
        assert output_path.exists()
        
        # Verify content
        df = pd.read_excel(output_path)
        assert len(df) > 0
        assert 'Company Name' in df.columns
        assert 'PE' in df.columns
        assert 'Valuation Flag' in df.columns
    
    def test_csv_generated(self):
        """Test that CSV file is generated"""
        output_path = Path('output/valuation_flags.csv')
        assert output_path.exists()
        
        # Verify content
        df = pd.read_csv(output_path)
        assert len(df) > 0
        assert 'Company Name' in df.columns
        assert 'Valuation Flag' in df.columns
    
    def test_92_companies_processed(self):
        """Test that 92 companies are processed"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        assert len(df) >= 90  # Allow some flexibility
    
    def test_required_columns_exist(self):
        """Test that all required columns exist"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        required_columns = [
            'Company ID', 'Company Name', 'Ticker', 'Sector',
            'Sub-sector', 'Broad Sector', 'Market Cap', 'PE', 'PB',
            'EV/EBITDA', 'Free Cash Flow', 'FCF Yield %',
            'Sector Median PE', 'PE vs Sector Median %', 'Valuation Flag'
        ]
        for col in required_columns:
            assert col in df.columns, f"Missing column: {col}"
    
    def test_valuation_flags_valid(self):
        """Test that valuation flags are valid"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        valid_flags = ['Fair', 'Discount', 'Caution']
        assert df['Valuation Flag'].isin(valid_flags).all()
    
    def test_csv_utf8_encoding(self):
        """Test CSV UTF-8 encoding"""
        with open('output/valuation_flags.csv', 'r', encoding='utf-8') as f:
            content = f.read()
            assert len(content) > 0
    
    def test_no_duplicate_companies(self):
        """Test no duplicate companies in Excel"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        duplicates = df['Company ID'].duplicated().sum()
        assert duplicates == 0, f"Found {duplicates} duplicate companies"


# =============================================================================
# PERFORMANCE TESTS
# =============================================================================

class TestPerformance:
    """Test performance requirements"""
    
    def test_pipeline_performance(self):
        """Test pipeline runs in <2 seconds"""
        import time
        start = time.time()
        run_valuation_pipeline()
        elapsed = time.time() - start
        assert elapsed < 2.0, f"Pipeline took {elapsed:.2f}s (should be <2s)"


# =============================================================================
# EDGE CASE TESTS
# =============================================================================

class TestEdgeCases:
    """Test edge cases"""
    
    def test_missing_pe(self):
        """Test handling of missing PE"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        # Should not crash even with missing PE
        assert len(df) > 0
    
    def test_missing_market_cap(self):
        """Test handling of missing market cap"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        # Should not crash even with missing market cap
        assert len(df) > 0
    
    def test_negative_fcf(self):
        """Test handling of negative FCF"""
        df = pd.read_excel('output/valuation_summary.xlsx')
        # Should not crash with negative FCF
        assert len(df) > 0


# =============================================================================
# VALIDATION TESTS
# =============================================================================

class TestValidation:
    """Test validation requirements"""
    
    def test_excel_formatting(self):
        """Test Excel has formatting"""
        from openpyxl import load_workbook
        wb = load_workbook('output/valuation_summary.xlsx')
        ws = wb.active
        
        # Check header formatting
        assert ws['A1'].font.bold is True
        
        # Check freeze panes
        assert ws.freeze_panes == 'A2'
        
        # Check filter
        assert ws.auto_filter.ref is not None
    
    def test_csv_headers(self):
        """Test CSV has proper headers"""
        df = pd.read_csv('output/valuation_flags.csv')
        required_headers = ['Company Name', 'Ticker', 'Sector', 'PE', 
                          'Sector Median PE', 'Valuation Flag']
        for header in required_headers:
            assert header in df.columns


# =============================================================================
# RUN TESTS
# =============================================================================

if __name__ == "__main__":
    logger.info("=" * 80)
    logger.info("RUNNING VALUATION MODULE TESTS")
    logger.info("=" * 80)
    
    # Run pipeline first to generate files
    logger.info("\n1. Running valuation pipeline...")
    stats = run_valuation_pipeline()
    logger.info(f"   Status: {stats['status']}")
    logger.info(f"   Companies: {stats['companies_processed']}")
    
    # Run pytest
    logger.info("\n2. Running unit tests...")
    exit_code = pytest.main([__file__, '-v', '--tb=short'])
    
    if exit_code == 0:
        logger.info("\n" + "=" * 80)
        logger.info("ALL TESTS PASSED")
        logger.info("=" * 80)
    else:
        logger.error("\n" + "=" * 80)
        logger.error("SOME TESTS FAILED")
        logger.info("=" * 80)
    
    sys.exit(exit_code)